#!/usr/bin/env python3
"""
Budget / forecast source for the Babyshop dashboard.

Parses the "Rullande P&L" forecast workbook (Forecast v1.xlsx) into a compact
monthly JSON — one entry per P&L line, 12 months (Jan→Dec) + full year — and
writes it to Firestore doc `funnel_cache/<WORKSPACE>__budget-2026`, which the
dashboard reads via /api/budget.

The forecast is a *budget*, not live data: it changes when Finance revises the
plan, not daily. So the flow is two-step and container-safe:

    1.  Locally, once per revision:  python3 budget_source.py path/to/Forecast.xlsx
        → parses the xlsx, writes budget_2026.json (committed to the repo).
    2.  The daily Cloud Run job (bq_source.main) calls push_budget(), which reads
        the committed budget_2026.json and writes it to Firestore. No xlsx is
        needed in the container, and the doc self-heals on every refresh.

Only two P&L lines reconcile with the dashboard's Funnel/BigQuery data and are
tracked live against plan (revenue, marketing spend); the rest are served as a
context plan table. See ``LINES`` and the revenue calibration note below.
"""
from __future__ import annotations

import datetime
import json
import os

# Reuse the exact Firestore wiring the rest of the dashboard uses.
from bq_source import COLLECTION, FIRESTORE_PROJECT, TTL, WORKSPACE, _credentials

HERE = os.path.dirname(os.path.abspath(__file__))
BUDGET_JSON = os.path.join(HERE, "budget_2026.json")
YEAR = 2026

# ── P&L line map ──────────────────────────────────────────────────────────────
# key -> (row in the "Blad1" sheet, human label, group, whether it's a cost line)
# Rows verified against Forecast v1.xlsx (columns C..N = M1..M12 = Jan..Dec, O = FY).
LINES = [
    ("gross_sales", 8,  "Gross sales",          "sales",   False),
    ("returns",     10, "Sales returns",        "sales",   True),
    ("net_sales",   13, "Net sales",            "sales",   False),
    ("cogs",        16, "Cost of goods sold",   "cogs",    True),
    ("gp1",         18, "Gross profit 1",       "profit",  False),
    ("direct_var",  34, "Direct variable costs","cost",    True),
    ("gp2",         36, "Gross profit 2",       "profit",  False),
    ("marketing",   40, "Marketing costs",      "cost",    True),
    ("gp3",         43, "Gross profit 3",       "profit",  False),
    ("opex",        56, "Operating expenses",   "cost",    True),
    ("other_income",59, "Other income",         "other",   False),
    ("ebitda",      61, "EBITDA",               "profit",  False),
    ("interest_da", 71, "Interest, D&A",        "cost",    True),
    ("ebt",         73, "EBT",                  "profit",  False),
    ("eat",         77, "Earnings after tax",   "profit",  False),
]
MONTH_STATUS_ROW = 5   # C5..N5 = "Actual" / "Forecast"
MONTH_COLS = list(range(3, 15))   # C(3) .. N(14)  → M1..M12
FY_COL = 15                       # O


def _num(v):
    if v is None:
        return None
    try:
        return round(float(v))
    except (TypeError, ValueError):
        return None


def parse_xlsx(path: str) -> dict:
    """Parse the forecast workbook into the budget dict."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Blad1"]

    month_status = [ws.cell(row=MONTH_STATUS_ROW, column=c).value for c in MONTH_COLS]
    month_status = [(s or "").strip() or "Forecast" for s in month_status]

    lines = {}
    for key, row, label, group, is_cost in LINES:
        monthly = [_num(ws.cell(row=row, column=c).value) for c in MONTH_COLS]
        fy = _num(ws.cell(row=row, column=FY_COL).value)
        # Store costs as positive magnitudes (the sheet keeps them negative);
        # `is_cost` tells the UI how to render them.
        if is_cost:
            monthly = [(-v if v is not None else None) for v in monthly]
            fy = (-fy if fy is not None else None)
        lines[key] = {
            "label": label, "group": group, "is_cost": is_cost,
            "monthly": monthly, "fy": fy,
        }

    return {
        "year": YEAR,
        "currency": "SEK",
        "month_status": month_status,          # 12 × "Actual" | "Forecast"
        "line_order": [k for k, *_ in LINES],
        "lines": lines,
        "source_file": os.path.basename(path),
        "note": "Rullande P&L forecast. Costs stored as positive magnitudes.",
    }


def load_budget_json() -> dict:
    with open(BUDGET_JSON, encoding="utf-8") as f:
        return json.load(f)


def push_budget(budget: dict | None = None) -> None:
    """Write the budget doc to Firestore (called by the daily job)."""
    from google.cloud import firestore
    budget = budget or load_budget_json()
    db = firestore.Client(project=FIRESTORE_PROJECT, credentials=_credentials())
    now = datetime.datetime.now(datetime.timezone.utc).timestamp()
    db.collection(COLLECTION).document(f"{WORKSPACE}__budget-2026").set({
        "data": budget,
        "fetched_at": firestore.SERVER_TIMESTAMP,
        "expires_at": now + TTL,
        "ttl_seconds": TTL,
        "workspace": WORKSPACE,
    })
    print(f"✓ budget-2026 written to Firestore ({len(budget['lines'])} lines)")


if __name__ == "__main__":
    import sys
    src = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
        "~/Downloads/Forecast v1 (1).xlsx")
    budget = parse_xlsx(src)
    with open(BUDGET_JSON, "w", encoding="utf-8") as f:
        json.dump(budget, f, ensure_ascii=False, indent=2)
    print(f"✓ parsed {src}")
    print(f"✓ wrote {BUDGET_JSON}")
    # Quick sanity dump.
    for k in ("gross_sales", "net_sales", "marketing", "gp3", "ebitda"):
        ln = budget["lines"][k]
        print(f"  {ln['label']:<22} Jan={ln['monthly'][0]:>13,} FY={ln['fy']:>14,}")
    # Push to Firestore too when creds are available (local convenience).
    if "--no-push" not in sys.argv:
        try:
            push_budget(budget)
        except Exception as e:  # noqa: BLE001
            print(f"⚠️  Firestore push skipped ({e!r}). JSON is written; the daily "
                  f"job will push it.")
